#!/usr/bin/env python3
"""
Pattern Matcher for 5Five Cricket Data
Matches patterns based on:
1. Over card sequence (same cards in same order)
2. First innings match (exact same 1st innings card sequence)
3. Final score match (exact same final scores for both teams)
4. Full Match Card Sequence (exact same cards for entire match)

Outputs an Excel file identical to input CSV plus analysis columns.
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime

class FiveCricketPatternMatcher:
    """Pattern Matcher for 5Five Cricket game"""
    
    def __init__(self, csv_file, excel_file=None):
        self.csv_file = csv_file
        self.excel_file = excel_file or "5five_cricket_patterns.xlsx"
        self.df = None
        
    def load_data(self):
        """Load CSV data"""
        if not os.path.exists(self.csv_file):
            print(f"File not found: {self.csv_file}")
            return False
        
        try:
            self.df = pd.read_csv(self.csv_file)
            print(f"✓ Loaded {len(self.df)} rows")
            return True
        except Exception as e:
            print(f"Error loading CSV: {e}")
            return False

    def analyze_patterns(self):
        """Analyze data and add pattern columns"""
        if self.df is None or len(self.df) == 0:
            print("⚠️  No data to analyze (empty dataframe)")
            return

        print("🔍 Analyzing patterns...")
        
        # Helper to get card from row
        def get_card(row):
            return str(row.get('Card', row.get('Runs', ''))).strip()

        # Group data by matches
        if 'Timestamp' in self.df.columns:
            self.df = self.df.sort_values(['Timestamp', 'Ball_Number'])
        
        # We need Round_ID
        round_id_col = 'Round_ID' if 'Round_ID' in self.df.columns else 'Match_ID'
        if round_id_col not in self.df.columns:
            print("No Round_ID column found")
            return

        matches_grouped = self.df.groupby(round_id_col, sort=False)
        
        if len(matches_grouped) == 0:
            print("⚠️  No matches found in data")
            return
        
        # HISTORY DATABASES
        # Over Sequence -> List of (Timestamp, MatchID_Inn_OverNum)
        hist_overs = {}
        # 1st Inn Sequence -> List of (Timestamp, MatchID)
        hist_inn1 = {}
        # Final Scores -> List of (Timestamp, MatchID)
        hist_scores = {}
        # Full Match Sequence -> List of (Timestamp, MatchID)
        hist_full = {}
        
        # First Pass: Build History
        print("   Building history...")
        
        match_meta = {} # RoundID -> {timestamp, final_score, inn1_seq}
        
        for round_id, match_rows in matches_grouped:
            match_rows = match_rows.sort_values('Ball_Number')
            start_time = match_rows.iloc[0]['Timestamp'] if 'Timestamp' in match_rows.columns else datetime.now().isoformat()
            
            # Identify Innings Split
            inn2_start_idx = -1
            for idx, row in match_rows.iterrows():
                 t2_ball = row.get('Team2_Ball', 0)
                 t2_over = row.get('Team2_Over', 0)
                 try:
                     if (pd.notna(t2_ball) and int(float(t2_ball)) > 0) or (pd.notna(t2_over) and int(float(t2_over)) > 0):
                         inn2_start_idx = idx
                         break
                 except: pass
            
            # Split rows
            loc_idx = 0
            if inn2_start_idx != -1:
                try:
                    loc_idx = match_rows.index.get_loc(inn2_start_idx)
                except: loc_idx = 0 # Fallback
            
            inn1_rows = match_rows.iloc[:loc_idx] if inn2_start_idx != -1 else match_rows
            inn2_rows = match_rows.iloc[loc_idx:] if inn2_start_idx != -1 else pd.DataFrame()
            
            # 1. Extract 1st Innings Sequence
            inn1_cards = []
            for _, r in inn1_rows.iterrows():
                c = get_card(r)
                if c and c != 'nan': inn1_cards.append(c)
            
            inn1_seq = ",".join(inn1_cards)
            if inn1_seq:
                if inn1_seq not in hist_inn1: hist_inn1[inn1_seq] = []
                hist_inn1[inn1_seq].append((start_time, round_id))
            
            # 2. Extract Overs (Inn 1)
            for i in range(0, len(inn1_cards), 6):
                chunk = inn1_cards[i:i+6]
                if len(chunk) == 6:
                    seq = ",".join(chunk)
                    over_id = f"{round_id}_1_{i//6 + 1}"
                    if seq not in hist_overs: hist_overs[seq] = []
                    hist_overs[seq].append((start_time, over_id))

            # 3. Extract Overs (Inn 2)
            inn2_cards = []
            for _, r in inn2_rows.iterrows():
                c = get_card(r)
                if c and c != 'nan': inn2_cards.append(c)
                
            for i in range(0, len(inn2_cards), 6):
                chunk = inn2_cards[i:i+6]
                if len(chunk) == 6:
                    seq = ",".join(chunk)
                    over_id = f"{round_id}_2_{i//6 + 1}"
                    if seq not in hist_overs: hist_overs[seq] = []
                    hist_overs[seq].append((start_time, over_id))
                    
            # 4. Final Scores
            last_row = match_rows.iloc[-1]
            s1 = str(last_row.get('Team1_Score', ''))
            s2 = str(last_row.get('Team2_Score', ''))
            
            def parse_score(s):
                if not isinstance(s, str): return 0
                try:
                    main_part = s.split(' ')[0]
                    runs = main_part.split('-')[0].split('/')[0]
                    return int(runs)
                except: return 0
            
            score_tuple = (parse_score(s1), parse_score(s2))
            if score_tuple not in hist_scores: hist_scores[score_tuple] = []
            hist_scores[score_tuple].append((start_time, round_id))
            
            # 5. Full Match Sequence
            full_seq = inn1_seq + "|" + ",".join(inn2_cards)
            if len(inn1_cards) + len(inn2_cards) > 0:
                 if full_seq not in hist_full: hist_full[full_seq] = []
                 hist_full[full_seq].append((start_time, round_id))
            
            match_meta[round_id] = {
                'start_time': start_time,
                'inn1_seq': inn1_seq,
                'final_score': score_tuple,
                'full_seq': full_seq,
                'inn2_start_idx': inn2_start_idx
            }

        # Second Pass: Populate Columns
        print("   Populating analysis columns...")
        
        # Ensure cols exist
        new_cols = [
            'Pattern_Over_Count', 'Pattern_Over_Last_Occur',
            'Pattern_1stInn_Count', 'Pattern_1stInn_Last_Occur',
            'Pattern_FinalScore_Count', 'Pattern_FinalScore_Last_Occur',
            'Pattern_Match_Card_Count', 'Pattern_Match_Card_Last_Occur'
        ]
        for c in new_cols:
             if c not in self.df.columns: self.df[c] = None
        
        for round_id, match_rows in matches_grouped:
            meta = match_meta.get(round_id)
            if not meta: continue
            
            start_time = meta['start_time']
            
            # Match Level Analysis
            
            # 1. Final Score
            fs_matches = [m for m in hist_scores.get(meta['final_score'], []) if m[0] < start_time]
            fs_count = len(fs_matches)
            fs_last = fs_matches[-1][1] if fs_matches else "None"
            
            # 2. 1st Innings
            inn1_matches = [m for m in hist_inn1.get(meta['inn1_seq'], []) if m[0] < start_time]
            inn1_count = len(inn1_matches)
            inn1_last = inn1_matches[-1][1] if inn1_matches else "None"
            
            # 3. Full Match Card Sequence
            fm_matches = [m for m in hist_full.get(meta['full_seq'], []) if m[0] < start_time]
            fm_count = len(fm_matches)
            fm_last = fm_matches[-1][1] if fm_matches else "None"
            
            # Iterate rows to fill
            inn2_start = meta['inn2_start_idx']
            
            curr_cards = []
            
            for idx, row in match_rows.iterrows():
                # Fill Match Level Stats
                self.df.at[idx, 'Pattern_FinalScore_Count'] = fs_count
                self.df.at[idx, 'Pattern_FinalScore_Last_Occur'] = fs_last
                
                self.df.at[idx, 'Pattern_Match_Card_Count'] = fm_count
                self.df.at[idx, 'Pattern_Match_Card_Last_Occur'] = fm_last
                
                # Innings specific
                is_inn2 = False
                if inn2_start != -1 and idx >= inn2_start:
                    is_inn2 = True
                    
                if is_inn2:
                    self.df.at[idx, 'Pattern_1stInn_Count'] = inn1_count
                    self.df.at[idx, 'Pattern_1stInn_Last_Occur'] = inn1_last
                else:
                    self.df.at[idx, 'Pattern_1stInn_Count'] = 0
                    self.df.at[idx, 'Pattern_1stInn_Last_Occur'] = "Current Inn1"
                
                # Over Analysis
                # Reset card buffer if new innings started this row
                if idx == inn2_start:
                    curr_cards = []
                
                c = get_card(row)
                if c and c != 'nan': curr_cards.append(c)
                
                # Default empty
                self.df.at[idx, 'Pattern_Over_Count'] = 0
                self.df.at[idx, 'Pattern_Over_Last_Occur'] = "None"
                
                # Check for completed over
                if len(curr_cards) > 0 and len(curr_cards) % 6 == 0:
                    # Completed an over
                    over_seq = ",".join(curr_cards[-6:])
                    over_idx = len(curr_cards) // 6
                    
                    matches = hist_overs.get(over_seq, [])
                    
                    # Filter
                    priors = []
                    curr_inn_num = 2 if is_inn2 else 1
                    
                    for t, oid in matches:
                        if t < start_time:
                            priors.append(oid)
                        elif t == start_time and oid.startswith(str(round_id)):
                            # Same match - check order
                            try:
                                parts = oid.split('_')
                                o_inn = int(parts[1])
                                o_num = int(parts[2])
                                
                                if o_inn < curr_inn_num:
                                    priors.append(oid)
                                elif o_inn == curr_inn_num and o_num < over_idx:
                                    priors.append(oid)
                            except: pass

                    # Assign to LAST 6 ROWS (the over just completed)
                    subset = match_rows.loc[:idx] 
                    valid_card_indices = []
                    for ri, r in subset.iloc[::-1].iterrows(): # Backwards
                        rc = get_card(r)
                        if rc and rc != 'nan':
                            valid_card_indices.append(ri)
                            if len(valid_card_indices) == 6: break
                    
                    o_count = len(priors)
                    o_last = priors[-1] if priors else "None"
                    
                    for vi in valid_card_indices:
                        self.df.at[vi, 'Pattern_Over_Count'] = o_count
                        self.df.at[vi, 'Pattern_Over_Last_Occur'] = o_last


    def save_excel(self):
        """Save to Excel with formatting"""
        print(f"💾 Saving to {self.excel_file}...")
        try:
            # Reorder columns: Original + New
            # Ensure new columns are at end
            cols = list(self.df.columns)
            # Remove new cols from list if they exist (to re-append at end)
            analy_cols = [
                'Pattern_Over_Count', 'Pattern_Over_Last_Occur',
                'Pattern_1stInn_Count', 'Pattern_1stInn_Last_Occur',
                'Pattern_FinalScore_Count', 'Pattern_FinalScore_Last_Occur',
                'Pattern_Match_Card_Count', 'Pattern_Match_Card_Last_Occur'
            ]
            
            base_cols = [c for c in cols if c not in analy_cols]
            final_cols = base_cols + analy_cols
            
            self.df = self.df[final_cols]
            
            self.df.to_excel(self.excel_file, index=False)
            
            # Auto-width formatting
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Header Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Column Widths
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                
                adj_width = min(max_length + 2, 50) # Cap width
                ws.column_dimensions[col_letter].width = adj_width
                
            wb.save(self.excel_file)
            print(f"✓ Report saved: {self.excel_file}")
            return True
            
        except Exception as e:
            print(f"Error saving Excel: {e}")
            return False

if __name__ == "__main__":
    import sys
    csv_path = "cricket_data.csv"
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
        
    matcher = FiveCricketPatternMatcher(csv_path)
    if matcher.load_data():
        matcher.analyze_patterns()
        matcher.save_excel()